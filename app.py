
# -*- coding: utf-8 -*-
"""
Plan Financier Stratégique Familial - Streamlit (v2)
- Page "Paramètres avancés" (mapping 50/30/20, quadrants, seuils Baby Steps)
- Éditeur de projets (CRUD) via st.data_editor
- Export Excel formaté (KPI + Graphiques) via openpyxl
"""
import io
import json
from datetime import datetime
from typing import Dict, Any

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st

st.set_page_config(page_title="Plan Financier Stratégique Familial",
                   page_icon="💡", layout="wide")

# ==============================
# Defaults
# ==============================
DEFAULT_503020 = {
    "Besoins": ["Scolarité", "Santé", "Loyer", "Transports essentiels", "Nourriture"],
    "Envies": ["Voyage", "Cadeaux", "Sorties", "Culture", "Électronique"],
    "Épargne/Dette": ["Épargne projet", "Fonds d'urgence", "Remboursement dette"]
}
DEFAULT_QUADRANT_SOURCES = {
    "E": ["Salaire William", "Job", "Salariat"],
    "S": ["Freelance", "Consulting"],
    "B": ["IIBA", "Atekys", "Entreprise familiale"],
    "I": ["Dividendes", "Intérêts", "Immobilier", "Rente"]
}
DEFAULT_CONFIG = {
    "currency": "FCFA",
    "emergency_target": 1_000_000,
    "months_ef": 3,
    "independence_target_ratio": 1.0,
    "rule_503020": DEFAULT_503020,
    "quadrants": DEFAULT_QUADRANT_SOURCES,
    "baby_step_min_ef_amount": 1_000_000
}

# ==============================
# Loaders
# ==============================
@st.cache_data
def load_default_projects() -> pd.DataFrame:
    df = pd.read_csv("projects_sample.csv", parse_dates=["Date_echeance"])
    return df

@st.cache_data
def load_default_transactions() -> pd.DataFrame:
    df = pd.read_csv("transactions_sample.csv", parse_dates=["Date"])
    return df

def ensure_numeric(df: pd.DataFrame, cols):
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    return df

def ensure_dates(df: pd.DataFrame, cols):
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")
    return df

# ==============================
# Session bootstrap
# ==============================
if "config" not in st.session_state:
    st.session_state["config"] = DEFAULT_CONFIG.copy()

if "projects_df" not in st.session_state:
    st.session_state["projects_df"] = load_default_projects()

if "tx_df" not in st.session_state:
    st.session_state["tx_df"] = load_default_transactions()

# ==============================
# Sidebar Navigation
# ==============================
st.sidebar.title("📚 Navigation")
page = st.sidebar.radio("Aller à :", ["Tableau de bord", "Projets (CRUD)", "Paramètres avancés", "Export"])

# Uploaders (disponibles partout via sidebar)
st.sidebar.markdown("---")
st.sidebar.subheader("📥 Import")
u_proj = st.sidebar.file_uploader("Importer Projets (CSV)", type=["csv"])
u_tx = st.sidebar.file_uploader("Importer Transactions (CSV)", type=["csv"])
if u_proj:
    st.session_state["projects_df"] = pd.read_csv(u_proj, parse_dates=["Date_echeance"])
if u_tx:
    st.session_state["tx_df"] = pd.read_csv(u_tx, parse_dates=["Date"])

# ==============================
# Helpers
# ==============================
def compute_kpis(projects_df: pd.DataFrame, tx_df: pd.DataFrame, config: Dict[str, Any]):
    tx = tx_df.copy()
    prj = projects_df.copy()

    # Types
    prj = ensure_numeric(prj, ["Budget_prevu","Budget_cotise","ROI_estime_pct"])
    prj = ensure_dates(prj, ["Date_echeance"])

    tx["Montant"] = pd.to_numeric(tx["Montant"], errors="coerce").fillna(0.0)
    tx["Date"] = pd.to_datetime(tx["Date"], errors="coerce")

    # Nature
    if "Nature" not in tx.columns:
        tx["Nature"] = np.where(tx["Montant"]>=0, "Revenu", "Dépense")
    if "Categorie" not in tx.columns:
        tx["Categorie"] = ""

    # 50/30/20 labels
    r503020 = config.get("rule_503020", DEFAULT_503020)
    def label_503020(cat: str) -> str:
        c = (cat or "").lower()
        if c in [x.lower() for x in r503020.get("Besoins",[])]:
            return "Besoins"
        if c in [x.lower() for x in r503020.get("Envies",[])]:
            return "Envies"
        if c in [x.lower() for x in r503020.get("Épargne/Dette",[])]:
            return "Épargne/Dette"
        return "Autre"
    tx["Label_503020"] = tx["Categorie"].apply(label_503020)

    # Revenu passif
    def is_passive(src: str) -> bool:
        src = (src or "").lower()
        passive_keywords = ["iiba", "dividende", "intérêt", "immobilier", "rente", "location", "invest"]
        return any(k in src for k in passive_keywords)
    tx["Revenu_Passif"] = np.where((tx["Nature"]=="Revenu") & (tx["Source"].astype(str).apply(is_passive)), 1, 0)

    # KPI Cashflow
    tx["Mois"] = tx["Date"].dt.to_period("M").astype(str)
    rev = tx[tx["Nature"]=="Revenu"].groupby("Mois")["Montant"].sum().rename("Revenus")
    dep = -tx[tx["Nature"]=="Dépense"].groupby("Mois")["Montant"].sum().rename("Dépenses")
    kpi_df = pd.concat([rev, dep], axis=1).fillna(0.0)
    kpi_df["Solde"] = kpi_df["Revenus"] - kpi_df["Dépenses"]

    revenus_cum = kpi_df["Revenus"].sum()
    depenses_cum = kpi_df["Dépenses"].sum()
    solde_cum = revenus_cum - depenses_cum

    # Emergency fund: réalisé = flux d'épargne/dette positifs
    ef_realise = tx[(tx["Nature"]=="Revenu") & (tx["Categorie"].str.lower().str.contains("urgence|épargne", na=False))]["Montant"].sum()
    avg_monthly_exp = kpi_df["Dépenses"].mean() if len(kpi_df)>0 else 0.0
    ef_required = max(config.get("emergency_target", 1_000_000),
                      config.get("months_ef", 3) * avg_monthly_exp)
    ef_coverage = float(ef_realise) / ef_required if ef_required else 0.0

    passive_income = tx[(tx["Nature"]=="Revenu") & (tx["Revenu_Passif"]==1)]["Montant"].sum()
    passive_ratio = float(passive_income) / revenus_cum if revenus_cum else 0.0

    independence_target = config.get("independence_target_ratio", 1.0)
    independence_attained = (passive_income >= depenses_cum) or (passive_ratio >= independence_target)

    # 50/30/20
    lab = tx[tx["Nature"]=="Dépense"].groupby("Label_503020")["Montant"].sum().abs()
    tot_dep = lab.sum()
    if tot_dep > 0:
        need_p = float(lab.get("Besoins",0))/tot_dep
        want_p = float(lab.get("Envies",0))/tot_dep
        save_p  = float(lab.get("Épargne/Dette",0))/tot_dep
        rule_ok = (need_p<=0.55) and (want_p<=0.35) and (save_p>=0.15)
    else:
        need_p = want_p = save_p = 0.0
        rule_ok = True

    # Phase
    if ef_coverage < 1.0 or solde_cum < 0:
        phase = "Stabilisation"
    elif passive_ratio < 0.5:
        phase = "Transition"
    else:
        phase = "Expansion"

    # Baby Step (simplifié)
    min_ef = config.get("baby_step_min_ef_amount", 1_000_000)
    if ef_realise < min_ef:
        baby_step = 1
    elif ef_realise < ef_required:
        baby_step = 3
    else:
        baby_step = 4

    return {
        "kpi_df": kpi_df, "revenus_cum": revenus_cum, "depenses_cum": depenses_cum, "solde_cum": solde_cum,
        "ef_realise": ef_realise, "ef_required": ef_required, "ef_coverage": ef_coverage,
        "passive_income": passive_income, "passive_ratio": passive_ratio,
        "independence_attained": independence_attained,
        "need_p": need_p, "want_p": want_p, "save_p": save_p, "rule_ok": rule_ok,
        "phase": phase, "baby_step": baby_step,
        "tx": tx, "projects": prj
    }

def quadrant_breakdown(tx_df: pd.DataFrame, quadrants: Dict[str, Any]) -> pd.DataFrame:
    revenus = tx_df[tx_df["Nature"]=="Revenu"].copy()
    revenus["Quadrant"] = "Autre"
    for quad, keys in quadrants.items():
        mask = False
        for k in keys:
            mask = mask | revenus["Source"].astype(str).str.contains(k, case=False, na=False)
        revenus.loc[mask, "Quadrant"] = quad
    return revenus.groupby("Quadrant")["Montant"].sum().reset_index()

def apply_council(prj: pd.DataFrame, ef_coverage: float, baby_step: int) -> pd.DataFrame:
    out = prj.copy()
    out["Avancement"] = out.apply(lambda r: (float(r["Budget_cotise"])/float(r["Budget_prevu"]) if float(r["Budget_prevu"]) else 0.0), axis=1)
    out["Actif_Passif"] = np.where(out["Type"].astype(str).str.lower().str_contains("actif", na=False), "Actif", "Passif") if hasattr(out["Type"].str, "str_contains") else np.where(out["Type"].astype(str).str.lower().str.contains("actif", na=False), "Actif", "Passif")

    def avis_kiyosaki(row) -> str:
        if row["Actif_Passif"] == "Actif":
            if float(row.get("ROI_estime_pct",0)) > 0:
                return "✅ Actif : à financer en priorité si ROI > coût. Utiliser revenus passifs / business."
            else:
                return "ℹ️ Actif neutre : préciser le ROI attendu avant d'engager de gros montants."
        else:
            return "⚠️ Passif : financer via revenus d'actifs ; si Fonds d'urgence < 100%, reporter."
    def avis_ramsey(row) -> str:
        if row["Actif_Passif"] == "Passif" and (baby_step < 3):
            return "⛔ Reporter (Baby Steps 1-3 : sécurité d'abord)."
        if str(row.get("Categorie","")).lower() in ["scolarité","santé"]:
            return "✅ Priorité vitale (OK même en Baby Steps 1-2)."
        if (ef_coverage < 1.0):
            return "🟠 Attendre couverture EF à 100%."
        return "✅ Autoriser avec budget mensuel fixe."
    def avis_orman(row) -> str:
        cat = str(row.get("Categorie","")).lower()
        if cat in ["scolarité","santé","administratif"]:
            return "✅ Personnes d'abord : à financer avant biens matériels."
        if row["Actif_Passif"] == "Passif":
            return "🟠 Passif : ne pas entamer EF ; vérifier 50/30/20."
        return "✅ Si contribue à la sécurité/indépendance, feu vert."
    out["Avis_Kiyosaki"] = out.apply(avis_kiyosaki, axis=1)
    out["Avis_Ramsey"]   = out.apply(avis_ramsey, axis=1)
    out["Avis_Orman"]    = out.apply(avis_orman, axis=1)
    return out

# ==============================
# PAGES
# ==============================
cfg = st.session_state["config"]
projects_df = st.session_state["projects_df"]
tx_df = st.session_state["tx_df"]

if page == "Tableau de bord":
    res = compute_kpis(projects_df, tx_df, cfg)

    st.title("🏠 Tableau de bord familial")
    col1, col2, col3, col4 = st.columns([1.2,1.2,1.2,1])
    with col1:
        st.metric("Revenus cumulés", f"{res['revenus_cum']:,.0f} FCFA".replace(",", " "))
    with col2:
        st.metric("Dépenses cumulées", f"{res['depenses_cum']:,.0f} FCFA".replace(",", " "))
    with col3:
        st.metric("Solde net", f"{res['solde_cum']:,.0f} FCFA".replace(",", " "))
    with col4:
        st.metric("Phase", res["phase"])
        st.metric("Baby Step (Ramsey)", f"{res['baby_step']}")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("Couverture Fonds d'urgence", f"{res['ef_coverage']*100:.0f}%", help=f"Réalisé {res['ef_realise']:,.0f} / Cible {res['ef_required']:,.0f} FCFA".replace(",", " "))
    with c2:
        st.metric("Ratio revenus passifs", f"{res['passive_ratio']*100:.0f}%")
    with c3:
        st.metric("Indépendance financière", "✅ Oui" if res["independence_attained"] else "❌ Non")
    with c4:
        st.metric("Règle 50/30/20", "✅ OK" if res["rule_ok"] else "⚠️ À surveiller",
                  help=f"Besoins {res['need_p']*100:.0f}% | Envies {res['want_p']*100:.0f}% | Épargne/Dette {res['save_p']*100:.0f}%")

    st.divider()
    ca, cb = st.columns(2)
    with ca:
        st.subheader("📈 Cashflow mensuel")
        if not res["kpi_df"].empty:
            cash = res["kpi_df"].reset_index()
            fig = px.line(cash, x="Mois", y=["Revenus","Dépenses","Solde"], markers=True)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Aucune transaction pour tracer le cashflow.")
    with cb:
        st.subheader("🍩 Répartition 50/30/20 (dépenses)")
        lab = tx_df[tx_df["Nature"]=="Dépense"].groupby("Label_503020")["Montant"].sum().abs()
        tot_dep = lab.sum()
        if tot_dep>0:
            donut_df = pd.DataFrame({
                "Catégorie":["Besoins","Envies","Épargne/Dette","Autre"],
                "Montant":[lab.get("Besoins",0), lab.get("Envies",0), lab.get("Épargne/Dette",0), lab.get("Autre",0)]
            })
            fig2 = px.pie(donut_df, values="Montant", names="Catégorie", hole=0.6)
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("Aucune dépense catégorisée pour la règle 50/30/20.")

    st.subheader("🧭 Quadrants Familiaux (Kiyosaki)")
    quad = quadrant_breakdown(res["tx"], cfg.get("quadrants", DEFAULT_QUADRANT_SOURCES))
    if not quad.empty:
        fig3 = px.bar(quad, x="Quadrant", y="Montant", text_auto=True)
        st.plotly_chart(fig3, use_container_width=True)
    else:
        st.info("Pas de revenus pour afficher les quadrants.")

    st.divider()
    st.subheader("📋 Projets : faisabilité & avis croisés")
    prj = apply_council(res["projects"], res["ef_coverage"], res["baby_step"])
    st.dataframe(prj[["Projet","Categorie","Type","Priorite","Budget_prevu","Budget_cotise","Avancement","Actif_Passif","Avis_Kiyosaki","Avis_Ramsey","Avis_Orman","Tags"]], use_container_width=True)

elif page == "Projets (CRUD)":
    st.title("🛠️ Éditeur de projets (CRUD)")
    st.caption("Ajoutez, modifiez ou marquez des lignes à supprimer (colonne 'Supprimer'). Cliquez ensuite sur 'Appliquer modifications'.")

    df_edit = st.session_state["projects_df"].copy()
    if "Supprimer" not in df_edit.columns:
        df_edit["Supprimer"] = False

    edited = st.data_editor(
        df_edit,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Budget_prevu": st.column_config.NumberColumn("Budget prévu", min_value=0, step=10_000),
            "Budget_cotise": st.column_config.NumberColumn("Budget côtisé", min_value=0, step=10_000),
            "ROI_estime_pct": st.column_config.NumberColumn("ROI estimé (%)", min_value=0, max_value=100, step=1),
            "Date_echeance": st.column_config.DateColumn("Échéance"),
            "Supprimer": st.column_config.CheckboxColumn("Supprimer")
        }
    )

    colA, colB = st.columns([1,1])
    with colA:
        if st.button("Appliquer modifications ✅"):
            cleaned = edited[edited["Supprimer"] != True].drop(columns=["Supprimer"])
            cleaned = ensure_numeric(cleaned, ["Budget_prevu","Budget_cotise","ROI_estime_pct"])
            cleaned = ensure_dates(cleaned, ["Date_echeance"])
            st.session_state["projects_df"] = cleaned
            st.success("Modifications enregistrées dans la session.")
    with colB:
        st.download_button("Télécharger Projets (CSV) ⬇️", data=edited.drop(columns=["Supprimer"]).to_csv(index=False).encode("utf-8"),
                           file_name="projets_edites.csv", mime="text/csv")

    st.divider()
    st.subheader("Transactions (lecture seule rapide)")
    st.dataframe(st.session_state["tx_df"].head(200), use_container_width=True)

elif page == "Paramètres avancés":
    st.title("⚙️ Paramètres avancés")
    config = st.session_state["config"].copy()

    st.subheader("Règle 50/30/20 — Mappage des catégories")
    col1, col2, col3 = st.columns(3)
    with col1:
        besoins = st.text_area("Besoins (séparés par virgules)", value=", ".join(config["rule_503020"].get("Besoins", [])))
    with col2:
        envies = st.text_area("Envies (séparés par virgules)", value=", ".join(config["rule_503020"].get("Envies", [])))
    with col3:
        epargne = st.text_area("Épargne/Dette (séparés par virgules)", value=", ".join(config["rule_503020"].get("Épargne/Dette", [])))

    st.subheader("Quadrants (sources de revenus) — JSON")
    quad_json = st.text_area("Quadrants JSON", value=json.dumps(config.get("quadrants", DEFAULT_QUADRANT_SOURCES), ensure_ascii=False, indent=2))

    st.subheader("Seuils — Fonds d'urgence & Indépendance")
    colA, colB, colC = st.columns(3)
    with colA:
        emergency_target = st.number_input("Objectif Fonds d'urgence (FCFA)", min_value=0, value=int(config.get("emergency_target",1_000_000)), step=100_000)
    with colB:
        months_ef = st.slider("Fonds d'urgence cible (mois de dépenses)", min_value=1, max_value=12, value=int(config.get("months_ef",3)))
    with colC:
        independence_target_ratio = st.slider("Cible indépendance : revenus passifs / dépenses", 0.0, 2.0, float(config.get("independence_target_ratio",1.0)), 0.1)

    st.subheader("Baby Steps (seuil initial minimal)")
    baby_min = st.number_input("Seuil minimal EF (FCFA) — déclenche Baby Step 1", min_value=0, value=int(config.get("baby_step_min_ef_amount",1_000_000)), step=100_000)

    csave, cload, cdl = st.columns([1,1,1])
    with csave:
        if st.button("Enregistrer paramètres ✅"):
            config["rule_503020"] = {
                "Besoins": [x.strip() for x in besoins.split(",") if x.strip()],
                "Envies": [x.strip() for x in envies.split(",") if x.strip()],
                "Épargne/Dette": [x.strip() for x in epargne.split(",") if x.strip()],
            }
            try:
                config["quadrants"] = json.loads(quad_json)
            except Exception:
                st.warning("JSON Quadrants invalide — paramètres précédents conservés.")
            config["emergency_target"] = emergency_target
            config["months_ef"] = months_ef
            config["independence_target_ratio"] = independence_target_ratio
            config["baby_step_min_ef_amount"] = baby_min
            st.session_state["config"] = config
            st.success("Paramètres enregistrés dans la session.")
    with cload:
        cfg_up = st.file_uploader("Charger config JSON", type=["json"])
        if cfg_up:
            try:
                cfg_loaded = json.load(cfg_up)
                st.session_state["config"] = cfg_loaded
                st.success("Configuration importée.")
            except Exception as e:
                st.error(f"Erreur lors du chargement : {e}")
    with cdl:
        st.download_button("Télécharger config JSON ⬇️",
                           data=json.dumps(st.session_state["config"], ensure_ascii=False, indent=2).encode("utf-8"),
                           file_name="config_familiale.json", mime="application/json")

elif page == "Export":
    st.title("📤 Export Excel formaté (KPI + Graphiques)")

    res = compute_kpis(projects_df, tx_df, cfg)
    prj_counsel = apply_council(res["projects"], res["ef_coverage"], res["baby_step"])

    st.write("Cet export génère un fichier .xlsx avec : feuilles **KPI**, **Cashflow**, **503020**, **Projets**, **Transactions** + graphiques (courbe Revenus/Dépenses et barres Solde).")

    if st.button("Générer l'export Excel ✅"):
        try:
            from openpyxl import Workbook
            from openpyxl.chart import LineChart, BarChart, Reference
            from openpyxl.styles import Font

            wb = Workbook()
            ws_kpi = wb.active
            ws_kpi.title = "KPI"

            # KPI sheet
            ws_kpi.append(["Indicateur", "Valeur"])
            ws_kpi["A1"].font = Font(bold=True)
            ws_kpi["B1"].font = Font(bold=True)
            ws_kpi.append(["Revenus cumulés", float(res["revenus_cum"])])
            ws_kpi.append(["Dépenses cumulées", float(res["depenses_cum"])])
            ws_kpi.append(["Solde net", float(res["solde_cum"])])
            ws_kpi.append(["EF réalisé", float(res["ef_realise"])])
            ws_kpi.append(["EF requis", float(res["ef_required"])])
            ws_kpi.append(["Couverture EF (%)", float(res["ef_coverage"]*100)])
            ws_kpi.append(["Revenus passifs", float(res["passive_income"])])
            ws_kpi.append(["Ratio passif (%)", float(res["passive_ratio"]*100)])
            ws_kpi.append(["Indépendance atteinte", "Oui" if res["independence_attained"] else "Non"])
            ws_kpi.append(["Phase", res["phase"]])
            ws_kpi.append(["Baby Step", int(res["baby_step"])])

            # Cashflow sheet
            ws_cf = wb.create_sheet("Cashflow")
            ws_cf.append(["Mois","Revenus","Dépenses","Solde"])
            for _, r in res["kpi_df"].reset_index().iterrows():
                ws_cf.append([r["Mois"], float(r["Revenus"]), float(r["Dépenses"]), float(r["Solde"])])

            # Charts on Cashflow
            if res["kpi_df"].shape[0] > 0:
                # Data ranges
                nrows = res["kpi_df"].shape[0] + 1  # include header
                # Line chart for Revenus/Dépenses
                line = LineChart()
                line.title = "Évolution Revenus vs Dépenses"
                data_line = Reference(ws_cf, min_col=2, min_row=1, max_col=3, max_row=nrows)
                cats_line = Reference(ws_cf, min_col=1, min_row=2, max_row=nrows)
                line.add_data(data_line, titles_from_data=True)
                line.set_categories(cats_line)
                ws_cf.add_chart(line, "F2")

                # Bar chart for Solde
                bar = BarChart()
                bar.title = "Solde net par mois"
                data_bar = Reference(ws_cf, min_col=4, min_row=1, max_row=nrows)
                cats_bar = Reference(ws_cf, min_col=1, min_row=2, max_row=nrows)
                bar.add_data(data_bar, titles_from_data=True)
                bar.set_categories(cats_bar)
                ws_cf.add_chart(bar, "F20")

            # 50/30/20 sheet
            ws_503020 = wb.create_sheet("503020")
            ws_503020.append(["Catégorie","Montant"])
            lab = res["tx"][res["tx"]["Nature"]=="Dépense"].groupby("Label_503020")["Montant"].sum().abs()
            for cat in ["Besoins","Envies","Épargne/Dette","Autre"]:
                ws_503020.append([cat, float(lab.get(cat, 0.0))])

            # Projets sheet (avec avis)
            ws_prj = wb.create_sheet("Projets")
            cols_prj = list(prj_counsel.columns)
            ws_prj.append(cols_prj)
            for _, r in prj_counsel.iterrows():
                row = []
                for c in cols_prj:
                    v = r.get(c, "")
                    if isinstance(v, (np.floating, np.integer)):
                        v = float(v)
                    row.append(v)
                ws_prj.append(row)

            # Transactions sheet
            ws_tx = wb.create_sheet("Transactions")
            cols_tx = list(res["tx"].columns)
            ws_tx.append(cols_tx)
            for _, r in res["tx"].iterrows():
                row = []
                for c in cols_tx:
                    v = r.get(c, "")
                    if isinstance(v, (np.floating, np.integer)):
                        v = float(v)
                    row.append(v)
                ws_tx.append(row)

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            st.download_button("Télécharger l'export Excel ⬇️",
                               data=output.getvalue(),
                               file_name=f"Export_Finances_Familiales_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.success("Export généré avec succès.")
        except Exception as e:
            st.error(f"Erreur lors de la génération : {e}")
